import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook



class   SRSCALCULATOR:

    def __init__(self,filepath):
        self.filepath = filepath

        # Load Excel and find the SRS sheet
        xls = pd.ExcelFile(self.filepath)
        self.sheet_name = next((name for name in xls.sheet_names if name.lower().startswith('srs')), None)
        if self.sheet_name is None:
            raise ValueError("No sheet starting with 'srs' found.")
        
        # Load both header rows
        self.df = pd.read_excel(self.filepath, header=[0, 1], sheet_name=self.sheet_name)

        # Extract first and second-level headers
        row0 = self.df.columns.get_level_values(0)
        row1 = self.df.columns.get_level_values(1)

        # Only keep the first set of "X. Something Score" columns
        self.df.columns = [str(col1).strip() for col1 in row1]  # Flatten to second row for scoring
        self.c = self.df.copy(deep=True)

        self.header_map = {}
        seen = set()

        for i, (head0, head1) in enumerate(zip(row0, row1)):
            header_str = str(head1).strip()
            match = re.match(r'^(\d+)\.', header_str)
            if match:
                q_num = int(match.group(1))
                if q_num not in seen:
                    self.header_map[q_num] = header_str
                    seen.add(q_num)

        print("🗺️ Header number mapping:", self.header_map)

    def sum_range(self, item_list):
        total = None  # Start with None so we can initialize with the first Series
        for num, func, fill in item_list:
            col = self.header_map.get(num)
            if col and col in self.c.columns:
                print(f"→ Applying {func.__name__} on: '{col}' | exists: {col in self.c.columns}")
                print(self.c[col].head(5))
                series = func(self.c[col], fill)
                if total is None:
                    total = series
                else:
                    total = total.add(series, fill_value=0)
        return total if total is not None else pd.Series([0] * len(self.df))
    
    def mod(self,args,i):
        a = args.copy()
        a = pd.Series(np.where(args.isin([1, 2, 3, 4]), args - 1, i), index=args.index)
        a = a.fillna(i)
        return a
    
    def reversemod(self,args,i):
        reverse_map = {1: 4, 2: 3, 3: 2, 4: 1}
        a = args.copy()
        mapped = a.map(reverse_map)
        mapped = mapped.fillna(a)        
        mapped = mapped - 1
        mapped = mapped.where(mapped <= 3, i)
        mapped = mapped.fillna(i)
        return pd.Series(mapped, index=a.index)
    
    def awr_total(self,a_series):
        def calculate(a):
            x = a
            x = 30 + 3*a 
            if a < 7:
                x -= 2
            elif a >= 8 or a <= 14:
                x -= 3
            else:
                x -= 4
            return x
        return a_series.apply(calculate)
    
    def cog_total(self,a_series):
        def calculate(a):
            x = a
            x = 35 + 2 * a
            if a == 1:
                return x
            for i, (low, high) in enumerate([(2,5), (6,9), (10,12), (13,16), (17,20), (21,24), (25,28), (29,31)], start=1):
                if low <= a <= high:
                    x -= i
                    break
            else:
                x -= 9
            return x
        return a_series.apply(calculate)
    
    def com_total(self,a_series):
        def calculate(a):
            x = a
            x = 35 + a 
            if a >= 4 or a <= 16:
                x += 1
            elif a >= 17 or a <= 30:
                x += 2
            elif a >= 31 or a <= 44:
                x += 3
            else:
                x += 4
            return x
        return a_series.apply(calculate)
    
    def mot_total(self,a_series):
        def calculate(a):
            x = a
            x = 37 + 2 * a
            if a > 10:
                x += 1
            return x
        return a_series.apply(calculate)
    
    def rrb_total(self,a_series):
        def calculate(a):
            x = a
            x = 40 + 2 * a
            return x
        return a_series.apply(calculate)
    
    def sci_total(self,a_series):
        def calculate(a):
            if a == 0:
                return 33
            elif a == 1:
                return 34
            elif a == 2:
                return 34  # subtractor = 1

            pattern = [3, 2, 2, 2, 2, 2]  # repeating pattern
            subtractor = 1
            a_index = a - 3
            pattern_index = 0

            while True:
                range_size = pattern[pattern_index % len(pattern)]
                if a_index < range_size:
                # within current range: subtractor increases within range
                    return 33 + a - (subtractor + a_index)
            # move to next range
                a_index -= range_size
                subtractor += range_size - 1  # subtractor increases within range, not at boundary
                pattern_index += 1
        return a_series.apply(calculate)
    
    
    def srs_total(self,a_series):
        def calculate(a):
            if a in [0, 1]:
                return 34
    
            score = 35
            a_index = a - 2  # Start offset
            pattern = [3, 2, 3, 2, 3, 3]  # repeating pattern unit
            i = 0

            while True:
                group_size = pattern[i % len(pattern)]
                if a_index < group_size:
                    return score
                a_index -= group_size
                score += 1
                i += 1
        return a_series.apply(calculate)
    
    def sum(self):
        c = self.c
        awr_items = [
            (2, self.mod, 0), (7, self.reversemod, 1), (25, self.mod, 1), (32, self.reversemod, 0),
            (45, self.reversemod, 1), (52, self.reversemod, 2), (54, self.mod, 0), (56, self.mod, 1)
        ]
        cog_items = [
            (5, self.mod, 1), (10, self.mod, 0), (15, self.reversemod, 1), (17, self.reversemod, 1),
            (30, self.mod, 0), (40, self.reversemod, 1), (42, self.mod, 0), (44, self.mod, 0),
            (48, self.reversemod, 1), (58, self.mod, 0), (59, self.mod, 0), (62, self.mod, 0)
        ]
        com_items = [
            (12, self.reversemod, 1), (13, self.mod, 0), (16, self.mod, 0), (18, self.mod, 0),
            (19, self.mod, 1), (21, self.reversemod, 1), (22, self.reversemod, 1), (26, self.reversemod, 1),
            (33, self.mod, 0), (35, self.mod, 0), (36, self.mod, 0), (37, self.mod, 0), (38, self.reversemod, 1),
            (41, self.mod, 0), (46, self.mod, 0), (47, self.mod, 0), (51, self.mod, 0), (53, self.mod, 0),
            (55, self.reversemod, 2), (57, self.mod, 0), (60, self.mod, 0), (61, self.mod, 0)
        ]
        mot_items = [
            (1, self.mod, 0), (3, self.reversemod, 1), (6, self.mod, 0), (9, self.mod, 1), (11, self.reversemod, 1),
            (23, self.mod, 0), (27, self.mod, 0), (34, self.mod, 0), (43, self.reversemod, 1), (64, self.mod, 0), (65, self.mod, 0)
        ]
        rrb_items = [
            (4, self.mod, 0), (8, self.mod, 0), (14, self.mod, 0), (20, self.mod, 0), (24, self.mod, 0),
            (28, self.mod, 1), (29, self.mod, 0), (31, self.mod, 1), (39, self.mod, 0),
            (49, self.mod, 0), (50, self.mod, 0), (63, self.mod, 0)
        ]
        c['awr'] = self.sum_range(awr_items)
        c['cog'] = self.sum_range(cog_items)
        c['com'] = self.sum_range(com_items)
        c['mot'] = self.sum_range(mot_items)
        c['rrb'] = self.sum_range(rrb_items)
        c['sci'] = c['awr'] + c['cog'] + c['com'] + c['mot']
        c['srs'] = c['awr'] + c['cog'] + c['com'] + c['mot'] + c['rrb']

        parentreport_1_32 = [
            (1, self.mod, 0),
            (2, self.mod, 0),
            (3, self.reversemod, 1),
            (4, self.mod, 0),
            (5, self.mod, 1),
            (6, self.mod, 0),
            (7, self.mod, 1),
            (8, self.mod, 0),
            (9, self.mod, 1),
            (10, self.mod, 0),
            (11, self.reversemod, 1),
            (12, self.reversemod, 1),
            (13, self.mod, 0),
            (14, self.mod, 0),
            (15, self.reversemod, 1),
            (16, self.mod, 0),
            (17, self.reversemod, 1),
            (18, self.mod, 0),
            (19, self.mod, 1),
            (20, self.mod, 0),
            (21, self.reversemod, 1),
            (22, self.reversemod, 1),
            (23, self.mod, 0),
            (24, self.mod, 0),
            (25, self.mod, 1),
            (26, self.reversemod, 1),
            (27, self.mod, 0),
            (28, self.mod, 1),
            (29, self.mod, 0),
            (30, self.mod, 0),
            (31, self.mod, 1),
            (32, self.reversemod, 0)
        ]
        parentreport_33_65 = [
            (33, self.mod, 0),
            (34, self.mod, 0),
            (35, self.mod, 0),
            (36, self.mod, 0),
            (37, self.mod, 0),
            (38, self.reversemod, 1),
            (39, self.mod, 0),
            (40, self.reversemod, 1),
            (41, self.mod, 0),
            (42, self.mod, 0),
            (43, self.mod, 0),
            (44, self.mod, 0),
            (45, self.reversemod, 1),
            (46, self.mod, 0),
            (47, self.mod, 0),
            (48, self.reversemod, 1),
            (49, self.mod, 0),
            (50, self.mod, 0),
            (51, self.mod, 0),
            (52, self.reversemod, 2),
            (53, self.mod, 0),
            (54, self.mod, 0),
            (55, self.reversemod, 2),
            (56, self.mod, 1),
            (57, self.mod, 0),
            (58, self.mod, 0),
            (59, self.mod, 0),
            (60, self.mod, 0),
            (61, self.mod, 0),
            (62, self.mod, 0),
            (63, self.mod, 0),
            (64, self.mod, 0),
            (65, self.mod, 0)
        ]
        c['awr_t'] = self.awr_total(c['awr'])
        c['cog_t'] = self.cog_total(c['cog'])
        c['com_t'] = self.com_total(c['com'])
        c['mot_t'] = self.mot_total(c['mot'])
        c['rrb_t'] = self.rrb_total(c['rrb'])
        c['sci_t'] = self.sci_total(c['sci'])
        c['srs_t'] = self.srs_total(c['srs'])

        c['sum1-32'] = self.sum_range(parentreport_1_32)
        c['sum33-65'] = self.sum_range(parentreport_33_65)
        return c['awr'],c['cog'],c['com'],c['mot'],c['rrb'],c['sci'],c['srs'],c['sum1-32'],c['sum33-65'],c['awr_t'],c['cog_t'],c['com_t'],c['mot_t'],c['rrb_t'],c['srs_t'],c['sci_t']



    def write_to_file(self):
        from openpyxl.utils import get_column_letter
        workbook = load_workbook(self.filepath)
        sheet = workbook[self.sheet_name]

        print("AWR:")
        print(self.sum_range([
            (2, self.mod, 0), (7, self.reversemod, 1), (25, self.mod, 1)
        ]))
        print("✅ AWR values:")
        print(self.c['awr'].head(10))

            # Identify header row (adjust if needed)
        header_row = 2 
        start_row = 3

            # Build a mapping from header name to Excel column letter
        header_map = {}
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=header_row, column=col).value
            if cell_value:
                header_map[cell_value.strip()] = get_column_letter(col)

            # Dictionary of header names in Excel to DataFrame columns
        column_mapping = {
            'Raw Score: Social Awareness': self.c['awr'],
            'Raw Score: Social Cognition': self.c['cog'],
            'Raw Score: Social Communication': self.c['com'],
            'Raw Score: Social Motivation': self.c['mot'],
            'Raw Score: Restricted Interest and Repetitive Behaviour': self.c['rrb'],
            'Raw Score: Autistic Mannerisms': self.c['rrb'],
            'Social Communication subscale raw score (Awr, Cog, Com, Mot)""': self.c['sci'],
            'Total raw score': self.c['srs'],
            'Sum of items 1-32': self.c['sum1-32'],
            'Sum of items 33-65': self.c['sum33-65'],
            'T Score: Social Awareness': self.c['awr_t'],
            'T Score: Social Cognition': self.c['cog_t'],
            'T Score: Social Communication': self.c['com_t'],
            'T Score: Social Motivation': self.c['mot_t'],
            'T Score: Autistic Mannerisms': self.c['rrb_t'],
            'T Score: Restricted Interest and Repetitive Behaviour': self.c['rrb_t'],
            'Social Communication subscale T score (Awr, Cog, Com, Mot)""': self.c['sci_t'],
            'Total T Score': self.c['srs_t']
        }

            # Dynamically write values using header mapping
        for header, series in column_mapping.items():
            col_letter = header_map.get(header)
            if not col_letter:
                print(f"⚠️ Warning: Header '{header}' not found in sheet.")
                continue
            for i, value in enumerate(series):
                sheet[f'{col_letter}{start_row + i}'] = value

            workbook.save(self.filepath)

       
   